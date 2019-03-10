"""Generates uneditable template with chosen schedule"""

#-*- coding: utf-8 -*-

import os
from calendar import Calendar
from flask import request, redirect, url_for
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import column_index_from_string as cifs
from app import MONTH_NAMES, WEEKDAY_NAMES
from app.schedules import show_schedule_helper


class Styles:
    """
    Styles for file
    """
    header_style = PatternFill(fgColor="b4bbc1", fill_type="solid")
    names_style = PatternFill(fgColor="c9ccce", fill_type="solid")
    footer_style = PatternFill(fgColor="55bbff", fill_type="solid")
    l4_style = PatternFill(fgColor="80D332", fill_type="solid")
    u_style = PatternFill(fgColor="FC33FF", fill_type="solid")
    thin_border_style = Border(left=Side(style="thin"),
                               right=Side(style="thin"),
                               top=Side(style="thin"),
                               bottom=Side(style="thin"))
    right_med_border_style = Border(left=Side(style="thin"),
                                    right=Side(style="medium"),
                                    top=Side(style="thin"),
                                    bottom=Side(style="thin"))
    right_bot_med_border_style = Border(left=Side(style="thin"),
                                        right=Side(style="medium"),
                                        top=Side(style="thin"),
                                        bottom=Side(style="medium"))
    bottom_med_border_style = Border(left=Side(style="thin"),
                                     right=Side(style="thin"),
                                     top=Side(style="thin"),
                                     bottom=Side(style="medium"))
    alignment_style = Alignment(horizontal="center",
                                vertical="center",
                                wrapText=True)


def draw_headers(ws, workplace, year, month_name):
    """
    Draws cells in file for workplace, year and name of the month
    """
    start_cell = ws.cell(row=1, column=1)
    i = 1
    while i <= 3:
        for val in [workplace, year, month_name]:
            cel = ws.cell(row=start_cell.row + i, column=2)
            cel.value = val
            cel.fill = Styles.header_style
            if i == 3:
                cel.border = Styles.bottom_med_border_style
            else:
                cel.border = Styles.thin_border_style
            i += 1

    ws.merge_cells(start_row=2, start_column=3, end_row=4, end_column=3)
    cel = ws.cell(row=start_cell.row + 1, column=cifs(start_cell.column) + 2)
    cel.value = "D"
    cel.fill = Styles.header_style
    cel.border = Styles.right_bot_med_border_style
    ws.cell(row=4, column=3).border = Styles.right_bot_med_border_style


def draw_inner_headers(ws, sched_dict):
    """
    Draws cells in file for workers names and inner headers
    """
    start = 0
    for worker in sched_dict["workers"]:
        ws.merge_cells(start_row=2, start_column=start + 4, end_row=3, end_column=start + 7)
        ws.cell(row=2, column=start + 4).value = worker
        ws.cell(row=2, column=start + 4).fill = Styles.names_style
        ws.cell(row=2, column=start + 4).border = Styles.right_med_border_style
        ws.cell(row=2, column=start + 7).border = Styles.right_med_border_style
        ws.cell(row=4, column=start + 4).value = "Od"
        ws.cell(row=4, column=start + 4).fill = Styles.header_style
        ws.cell(row=4, column=start + 4).border = Styles.bottom_med_border_style
        ws.cell(row=4, column=start + 5).value = "Do"
        ws.cell(row=4, column=start + 5).fill = Styles.header_style
        ws.cell(row=4, column=start + 5).border = Styles.bottom_med_border_style
        ws.cell(row=4, column=start + 6).value = "H"
        ws.cell(row=4, column=start + 6).fill = Styles.header_style
        ws.cell(row=4, column=start + 6).border = Styles.bottom_med_border_style
        ws.cell(row=4, column=start + 7).value = "Podpis"
        ws.cell(row=4, column=start + 7).fill = Styles.header_style
        ws.cell(row=4, column=start + 7).border = Styles.right_bot_med_border_style
        start += 4


def draw_data(ws, sched_dict):
    """
    Draws cells with everyday data in file
    """
    start = ws.cell(row=5, column=2)
    c_r = 0
    for day in Calendar().itermonthdays2(sched_dict["year"], sched_dict["month"]):
        c_c = 0
        if day[0] > 0:
            day_name_cel = ws.cell(row=start.row + c_r, column=cifs(start.column) + c_c)
            day_name_cel.value = WEEKDAY_NAMES[day[1]]
            day_name_cel.fill = Styles.header_style
            day_name_cel.border = Styles.thin_border_style

            day_number_cell = ws.cell(row=day_name_cel.row, column=cifs(day_name_cel.column) + 1)
            day_number_cell.value = day[0]
            day_number_cell.fill = Styles.header_style
            day_number_cell.border = Styles.right_med_border_style

            # hours and events
            for worker in sched_dict["workers"]:
                if WEEKDAY_NAMES[day[1]] not in ["Sobota", "Niedziela"]:
                    color = PatternFill(fgColor="ffffff", fill_type="solid")
                else:
                    color = PatternFill(fgColor="b4bbc1", fill_type="solid")
                event = sched_dict["event-%s-%d-%02d-%02d" % (worker, sched_dict["year"], sched_dict["month"],
                                                              day[0])]
                sum_val = sched_dict["sum-%s-%d-%02d-%02d" % (worker, sched_dict["year"], sched_dict["month"],
                                                              day[0])]
                if event in ["UW", "UNŻ", "UO", "UOJ", "UR"]:
                    color = Styles.u_style
                elif event == "L4":
                    color = Styles.l4_style

                cel = ws.cell(row=start.row + c_r, column=c_c + 4)
                if sum_val == 0:
                    cel.value = "--"
                else:
                    cel.value = sched_dict["begin-%s-%d-%02d-%02d" % (worker, sched_dict["year"], sched_dict["month"],
                                                                      day[0])]
                cel.fill = color
                cel.border = Styles.thin_border_style
                cel = ws.cell(row=start.row + c_r, column=c_c + 5)
                if sum_val == 0:
                    cel.value = "--"
                else:
                    cel.value = sched_dict["end-%s-%d-%02d-%02d" % (worker, sched_dict["year"], sched_dict["month"],
                                                                    day[0])]
                cel.border = Styles.thin_border_style
                cel.fill = color
                cel = ws.cell(row=start.row + c_r, column=c_c + 6)
                cel.value = sum_val
                cel.border = Styles.thin_border_style
                cel.fill = color
                cel = ws.cell(row=start.row + c_r, column=c_c + 7)

                if event == "off":
                    cel.value = "X"
                elif event == "in_work":
                    cel.value = "........"
                else:
                    cel.value = event

                cel.border = Styles.right_med_border_style
                cel.fill = color
                c_c += 4
            c_r += 1


def draw_footers(ws, sched_dict):
    """
    Draw footer cells in file
    """
    start = ws.cell(row=ws.max_row + 1, column=2)
    ws.merge_cells(start_row=start.row, start_column=2, end_row=start.row, end_column=cifs(start.column) + 1)
    start.value = "H/mc:"
    start.fill = Styles.footer_style
    start.border = Styles.right_med_border_style
    ws.cell(row=start.row, column=cifs(start.column) + 1).border = Styles.right_med_border_style
    start_col = 0
    for worker in sched_dict["workers"]:
        ws.merge_cells(start_row=start.row, start_column=start_col + 4, end_row=start.row,
                       end_column=start_col + 7)
        cel = ws.cell(row=start.row, column=start_col + 4)
        cel.value = "Suma H:"
        cel.fill = Styles.footer_style
        cel.border = Styles.right_med_border_style
        ws.cell(row=start.row, column=start_col + 7).border = Styles.right_med_border_style
        start_col += 4

    start = ws.cell(row=ws.max_row + 1, column=2)
    ws.merge_cells(start_row=start.row, start_column=2, end_row=start.row, end_column=cifs(start.column) + 1)
    start.value = sched_dict["hours"]
    start.fill = Styles.footer_style
    start.border = Styles.right_med_border_style
    ws.cell(row=start.row, column=cifs(start.column) + 1).border = Styles.right_med_border_style
    start_col = 0
    for worker in sched_dict["workers"]:
        ws.merge_cells(start_row=start.row, start_column=start_col + 4, end_row=start.row,
                       end_column=start_col + 7)
        cel = ws.cell(row=start.row, column=start_col + 4)
        cel.value = sched_dict["workers_hours"][worker]
        cel.border = Styles.right_med_border_style
        ws.cell(row=start.row, column=start_col + 7).border = Styles.right_med_border_style
        if cel.value > sched_dict["hours"]:
            cel.fill = PatternFill(fgColor="ff7878", fill_type="solid")
        elif cel.value < sched_dict["hours"]:
            cel.fill = PatternFill(fgColor="faaa01", fill_type="solid")
        start_col += 4

    start = ws.cell(row=ws.max_row + 1, column=2)
    ws.merge_cells(start_row=start.row, start_column=2, end_row=start.row, end_column=cifs(start.column) + 1)
    start.value = "Pozostało:"
    start.fill = Styles.footer_style
    start.border = Styles.right_med_border_style
    ws.cell(row=start.row, column=cifs(start.column) + 1).border = Styles.right_med_border_style
    start_col = 0
    for worker in sched_dict["workers"]:
        ws.merge_cells(start_row=start.row, start_column=start_col + 4, end_row=start.row,
                       end_column=start_col + 7)
        cel = ws.cell(row=start.row, column=start_col + 4)
        cel.value = sched_dict["hours"] - sched_dict["workers_hours"][worker]
        cel.border = Styles.right_med_border_style
        ws.cell(row=start.row, column=start_col + 7).border = Styles.right_med_border_style
        if int(cel.value) < 0:
            cel.fill = PatternFill(fgColor="ff7878", fill_type="solid")
        elif int(cel.value) > 1:
            cel.fill = PatternFill(fgColor="faaa01", fill_type="solid")
        start_col += 4

    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).alignment = Styles.alignment_style


def get_sched_data():
    """
    Creates dictionary with basic schedule data
    :return: dictionary with data
    """
    sched = request.args.get("schd")
    sched_v = request.args.get("v")
    sched_dict = show_schedule_helper.show_schedule_helper(sched, sched_v)
    return sched_dict


def save_file(wb, workplace, month, year, filename):
    """
    Saves file
    """

    path = "app/xlsx_files/download/%s/%s/%s" % (year, month, workplace.replace(" ", "_"))
    if not os.path.exists(path):
        os.makedirs(path)
    wb.save("%s/%s" % (path, filename))


def create_file():
    """
    Creates file
    :return:
    """
    sched_dict = get_sched_data()
    workplace = sched_dict["workplace"]
    year = sched_dict["year"]
    month = sched_dict["month"]
    month_name = MONTH_NAMES[month - 1]
    version = sched_dict["version"]

    wb = Workbook()
    ws = wb.active
    ws.title = "grafik_na_%s" % month_name

    draw_headers(ws, workplace, year, month_name)
    draw_inner_headers(ws, sched_dict)
    draw_data(ws, sched_dict)
    draw_footers(ws, sched_dict)
    filename = "%s_grafik_na_%02d_%s_v_%s.xlsx" % (workplace.replace(" ", "_"), month, year, version)
    save_file(wb, workplace, month, year, filename)

    return redirect(url_for("xlsx.download_schedule", year=year, month=month, workplace=workplace, filename=filename))