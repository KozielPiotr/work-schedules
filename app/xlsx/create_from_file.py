"""Creates schedule from file"""

#-*- coding: utf-8 -*-

from calendar import Calendar
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string as cifs
from flask import redirect, url_for, render_template, flash, request
from app import MONTH_NAMES, WEEKDAY_NAMES
from app.models import Shop, Billing_period
from app.schedules import prev_schedule


def wrong_file(val, month, year, workplace):
    """
    Protects from loading file with wrong data
    :param val: string to error message
    :param month: month
    :param year: year
    :param workplace: workplace
    """
    flash("W pliku nie znaleziono grafiku dla wybranego %s." % val)
    return redirect(url_for("xlsx.upload_file", month=month, year=year, workplace=workplace))


def create_sched():
    """
    Renders template with schedule
    :return: template with schedule
    """
    path = request.args.get("path")
    filename = request.args.get("filename")
    file_path = path + filename
    year = int(request.args.get("year"))
    month = int(request.args.get("month"))
    workplace = request.args.get("workplace")
    hours = request.args.get("hours")
    db_workplace = Shop.query.filter_by(shopname=workplace).first()

    month_name = MONTH_NAMES[month - 1]
    cal = Calendar()
    prev = prev_schedule.prev_schedule(month, year, workplace)
    wb = load_workbook(file_path, data_only=True)

    # finding sheet
    sheets = wb.sheetnames
    sh = None
    for sheet in sheets:
        sheet_month = sheet.lower().replace("grafik_", "").lower()
        if sheet_month.capitalize() == month_name:
            sh = wb[sheet]
    if sh is None:
        flash("W pliku nie znaleziono grafiku na wybrany miesiąc.")
        return redirect(url_for("xlsx.upload_file", month=month, year=year, workplace=workplace))

    # finding month cell
    month_cell = None
    for row in range(int(sh.max_row)):
        for col in range(int(sh.max_column)):
            current_cell = sh.cell(row=row + 1, column=col + 1)
            if type(current_cell.value) is str and current_cell.value.lower().replace(" ", "") == month_name.lower():
                month_cell = current_cell
    if month_cell is None:
        return wrong_file("miesiąca", month, year, workplace)

    # finding year
    year_cell = None
    for row in range(int(sh.max_row)):
        for col in range(int(sh.max_column)):
            current_cell = sh.cell(row=row + 1, column=col + 1)
            if type(current_cell.value) is int and current_cell.value == year:
                year_cell = current_cell
    if year_cell is None:
        return wrong_file("roku", month, year, workplace)

    # finding workplace
    workplace_cell = None
    for row in range(int(sh.max_row)):
        for col in range(int(sh.max_column)):
            current_cell = sh.cell(row=row + 1, column=col + 1)
            if type(current_cell.value) is str and current_cell.value.lower() == workplace.lower():
                workplace_cell = current_cell
    if workplace_cell is None:
        return wrong_file("sklepu", month, year, workplace)

    # finding workers
    workers = {}
    wrong_workers = []
    db_workers = []
    worker_row = workplace_cell.row
    worker_col = cifs(workplace_cell.column) + 2
    name_cell = sh.cell(row=worker_row, column=worker_col)
    surename_cell = sh.cell(row=worker_row + 1, column=worker_col)
    while name_cell.value is not None and surename_cell.value is not None:
        name = name_cell.value.replace(" ", "") + " " + surename_cell.value.replace(" ", "")
        workers[name] = surename_cell
        # workers.append(name)
        worker_col += 4
        name_cell = sh.cell(row=worker_row, column=worker_col)
        surename_cell = sh.cell(row=worker_row + 1, column=worker_col)
    for worker in db_workplace.works.all():
        db_workers.append(str(worker))

    workers_to_schd = []
    for worker in workers:
        workers_to_schd.append(str(worker.replace(" ", "_")))
    for worker in workers_to_schd:
        if worker.replace("_", " ") not in db_workers:
            wrong_workers.append(worker.replace("_", " "))
    if wrong_workers:
        flash("Sprawdź plik. Pracownicy nieprzypisani do sklepu: %s" % wrong_workers)
        return redirect(url_for("xlsx.upload_file", month=month, year=year, workplace=workplace))

    # building dictionary with data for each day for each worker
    shdict = {}
    events = ("off", "in_work", "UW", "UNŻ", "L4", "UO", "UOJ", "UR", "UB")
    for worker in workers:
        for day in cal.itermonthdays(year, month):
            if day > 0:

                # begin hour
                begin = sh.cell(row=workers[worker].row + 1 + day, column=cifs(workers[worker].column)).value
                if begin is None:
                    begin = ""
                shdict["begin-%s-%d-%02d-%02d" % (worker.replace(" ", "_"), year, month, day)] = begin

                # end hour
                end = sh.cell(row=workers[worker].row + 1 + day, column=cifs(workers[worker].column) + 1).value
                if end is None:
                    end = ""
                shdict["end-%s-%d-%02d-%02d" % (worker.replace(" ", "_"), year, month, day)] = end

                # event
                event = sh.cell(row=workers[worker].row + 1 + day, column=cifs(workers[worker].column) + 3).value
                if event is None:
                    event = "in_work"
                elif event == "X":
                    event = "off"
                if event not in events:
                    flash("Niewłaściwe oznaczenie zdarzenia w dniu %02d-%02d-%04d u pracownika %s." % (day, month, year,
                                                                                                       worker))
                    return redirect(url_for("xlsx.upload_file", month=month, year=year, workplace=workplace))
                shdict["event-%s-%d-%02d-%02d" % (worker.replace(" ", "_"), year, month, day)] = event

    flash("Wczytałem grafik na %s %s" % (month_name.lower(), year))

    return render_template("xlsx/empty_schedule.html", title="grafiki", workplace=workplace, year=year, month=month,
                           workers=workers_to_schd, month_name=month_name, wdn=WEEKDAY_NAMES, cal=cal,
                           Billing_period=Billing_period, shdict=shdict, hours=hours, prev_shdict=prev["prev_shdict"],
                           prev_month=prev["month"], prev_hours=prev["hours"], prev_month_name=prev["month_name"],
                           prev_year=prev["year"], prev_workers=prev["workers"], workers_hours=prev["workers_hours"])
