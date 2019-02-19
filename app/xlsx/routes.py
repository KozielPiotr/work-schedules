"""
Creates xlsx files based on accepted versions of schedule and loads xlsx files to create new schedules
- to_xlsx(): Creates xlsx file with schedule
- download_schedule(): Allows to download file with schedule
- upload_file(): Uploads chosen .xlsx file with schedule
"""

#-*- coding: utf-8 -*-

import os
import calendar
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import column_index_from_string as cifs
from flask import request, redirect, url_for, current_app, send_from_directory, render_template, flash, abort
from flask_login import login_required
from werkzeug.utils import secure_filename
from app import schedules
from app.xlsx import bp
from app.xlsx.forms import UploadFile
from app.models import User, Shop, Billing_period, Schedule, Personal_schedule


@bp.route("/xlsx", methods=["GET", "POST"])
@login_required
def to_xlsx():
    """
    Creates xlsx file with schedule
    :return: redirects to template with generated link to created xlsx file with schedule
    """
    cal = calendar.Calendar()

    # getting dict with schedule data
    sched = request.args.get("schd")
    sched_v = request.args.get("v")
    sched_dict = schedules.routes.show_schedule_helper(sched, sched_v)

    # preparing data for xlsx file
    workplace = sched_dict["workplace"]
    year = sched_dict["year"]
    month = sched_dict["month"]
    version = sched_dict["version"]

    month_names = ["Styczeń", "Luty", "Marzec", "Kwiecień", "Maj", "Czerwiec", "Lipiec", "Sierpień",
                   "Wrzesień", "Październik", "Listopad", "Grudzień"]
    weekday_names = ["Poniedziałek", "Wtorek", "Środa", "Czwartek", "Piątek", "Sobota", "Niedziela"]
    month_name = month_names[month-1]

    # starting work with file
    wb = Workbook()
    ws = wb.active
    ws.title = "grafik_na_%s" % month_name

    # style patterns for file
    header_style = PatternFill(fgColor="b4bbc1", fill_type="solid")
    names_style = PatternFill(fgColor="c9ccce", fill_type="solid")
    footer_style = PatternFill(fgColor="55bbff", fill_type="solid")
    l4_style = PatternFill(fgColor="80D332", fill_type="solid")
    u_style = PatternFill(fgColor="FC33FF", fill_type="solid")
    thin_border_style = Border(left=Side(style="thin"),
                               right=Side(style="thin"),
                               top=Side(style="thin"),
                               bottom=Side(style="thin"))
    right_med_border_style = Border(left=Side(style="thin"),
                                    right=Side(style="medium"),
                                    top=Side(style="thin"),
                                    bottom=Side(style="thin"))
    right_bot_med_border_style = Border(left=Side(style="thin"),
                                        right=Side(style="medium"),
                                        top=Side(style="thin"),
                                        bottom=Side(style="medium"))
    bottom_med_border_style = Border(left=Side(style="thin"),
                                     right=Side(style="thin"),
                                     top=Side(style="thin"),
                                     bottom=Side(style="medium"))

    alignment_style = Alignment(horizontal="center",
                                vertical="center",
                                wrapText=True)

    start_cell = ws.cell(row=1, column=1)

    # workplace, year and name of the month
    i = 1
    while i <= 3:
        for val in [workplace, year, month_name]:
            cel = ws.cell(row=start_cell.row+i, column=2)
            cel.value = val
            cel.fill = header_style
            if i == 3:
                cel.border = bottom_med_border_style
            else:
                cel.border = thin_border_style
            i += 1

    ws.merge_cells(start_row=2, start_column=3, end_row=4, end_column=3)
    cel = ws.cell(row=start_cell.row+1, column=cifs(start_cell.column)+2)
    cel.value = "D"
    cel.fill = header_style
    cel.border = right_bot_med_border_style
    ws.cell(row=4, column=3).border = right_bot_med_border_style

    # workers names and headers
    start = 0
    for worker in sched_dict["workers"]:
        ws.merge_cells(start_row=2, start_column=start+4, end_row=3, end_column=start+7)
        ws.cell(row=2, column=start+4).value = worker
        ws.cell(row=2, column=start+4).fill = names_style
        ws.cell(row=2, column=start+4).border = right_med_border_style
        ws.cell(row=2, column=start+7).border = right_med_border_style
        ws.cell(row=4, column=start+4).value = "Od"
        ws.cell(row=4, column=start+4).fill = header_style
        ws.cell(row=4, column=start+4).border = bottom_med_border_style
        ws.cell(row=4, column=start+5).value = "Do"
        ws.cell(row=4, column=start+5).fill = header_style
        ws.cell(row=4, column=start+5).border = bottom_med_border_style
        ws.cell(row=4, column=start+6).value = "H"
        ws.cell(row=4, column=start+6).fill = header_style
        ws.cell(row=4, column=start+6).border = bottom_med_border_style
        ws.cell(row=4, column=start+7).value = "Podpis"
        ws.cell(row=4, column=start+7).fill = header_style
        ws.cell(row=4, column=start+7).border = right_bot_med_border_style
        start += 4

    # days and data for each worker
    start = ws.cell(row=5, column=2)
    c_r = 0
    for day in cal.itermonthdays2(sched_dict["year"], sched_dict["month"]):
        c_c = 0
        if day[0] > 0:
            day_name_cel = ws.cell(row=start.row+c_r, column=cifs(start.column)+c_c)
            day_name_cel.value = weekday_names[day[1]]
            day_name_cel.fill = header_style
            day_name_cel.border = thin_border_style

            day_number_cell = ws.cell(row=day_name_cel.row, column=cifs(day_name_cel.column)+1)
            day_number_cell.value = day[0]
            day_number_cell.fill = header_style
            day_number_cell.border = right_med_border_style

            # hours and events
            for worker in sched_dict["workers"]:
                if weekday_names[day[1]] not in ["Sobota", "Niedziela"]:
                    color = PatternFill(fgColor="ffffff", fill_type="solid")
                else:
                    color = PatternFill(fgColor="b4bbc1", fill_type="solid")
                event = sched_dict["event-%s-%d-%02d-%02d" % (worker, sched_dict["year"], sched_dict["month"],
                                                              day[0])]
                sum_val = sched_dict["sum-%s-%d-%02d-%02d" % (worker, sched_dict["year"], sched_dict["month"],
                                                              day[0])]
                if event in ["UW", "UNŻ", "UO", "UOJ", "UR"]:
                    color = u_style
                elif event == "L4":
                    color = l4_style

                cel = ws.cell(row=start.row+c_r, column=c_c+4)
                if sum_val == 0:
                    cel.value = "--"
                else:
                    cel.value = sched_dict["begin-%s-%d-%02d-%02d" % (worker, sched_dict["year"], sched_dict["month"],
                                                                      day[0])]
                cel.fill = color
                cel.border = thin_border_style
                cel = ws.cell(row=start.row + c_r, column=c_c + 5)
                if sum_val == 0:
                    cel.value = "--"
                else:
                    cel.value = sched_dict["end-%s-%d-%02d-%02d" % (worker, sched_dict["year"], sched_dict["month"],
                                                                    day[0])]
                cel.border = thin_border_style
                cel.fill = color
                cel = ws.cell(row=start.row + c_r, column=c_c + 6)
                cel.value = sum_val
                cel.border = thin_border_style
                cel.fill = color
                cel = ws.cell(row=start.row + c_r, column=c_c + 7)

                if event == "off":
                    cel.value = "X"
                elif event == "in_work":
                    cel.value = "........"
                else:
                    cel.value = event

                cel.border = right_med_border_style
                cel.fill = color
                c_c += 4
            c_r += 1

    # footer with hours
    start = ws.cell(row=ws.max_row+1, column=2)
    ws.merge_cells(start_row=start.row, start_column=2, end_row=start.row, end_column=cifs(start.column)+1)
    start.value = "H/mc:"
    start.fill = footer_style
    start.border = right_med_border_style
    ws.cell(row=start.row, column=cifs(start.column)+1).border = right_med_border_style
    start_col = 0
    for worker in sched_dict["workers"]:
        ws.merge_cells(start_row=start.row, start_column=start_col+4, end_row=start.row,
                       end_column=start_col+7)
        cel = ws.cell(row=start.row, column=start_col+4)
        cel.value = "Suma H:"
        cel.fill = footer_style
        cel.border = right_med_border_style
        ws.cell(row=start.row, column=start_col+7).border = right_med_border_style
        start_col += 4

    start = ws.cell(row=ws.max_row + 1, column=2)
    ws.merge_cells(start_row=start.row, start_column=2, end_row=start.row, end_column=cifs(start.column)+1)
    start.value = sched_dict["hours"]
    start.fill = footer_style
    start.border = right_med_border_style
    ws.cell(row=start.row, column=cifs(start.column)+1).border = right_med_border_style
    start_col = 0
    for worker in sched_dict["workers"]:
        ws.merge_cells(start_row=start.row, start_column=start_col+4, end_row=start.row,
                       end_column=start_col+7)
        cel = ws.cell(row=start.row, column=start_col+4)
        cel.value = sched_dict["workers_hours"][worker]
        cel.border = right_med_border_style
        ws.cell(row=start.row, column=start_col+7).border = right_med_border_style
        if cel.value > sched_dict["hours"]:
            cel.fill = PatternFill(fgColor="ff7878", fill_type="solid")
        elif cel.value < sched_dict["hours"]:
            cel.fill = PatternFill(fgColor="faaa01", fill_type="solid")
        start_col += 4

    start = ws.cell(row=ws.max_row+1, column=2)
    ws.merge_cells(start_row=start.row, start_column=2, end_row=start.row, end_column=cifs(start.column)+1)
    start.value = "Pozostało:"
    start.fill = footer_style
    start.border = right_med_border_style
    ws.cell(row=start.row, column=cifs(start.column)+1).border = right_med_border_style
    start_col = 0
    for worker in sched_dict["workers"]:
        ws.merge_cells(start_row=start.row, start_column=start_col+4, end_row=start.row,
                       end_column=start_col+7)
        cel = ws.cell(row=start.row, column=start_col+4)
        cel.value = sched_dict["hours"] - sched_dict["workers_hours"][worker]
        cel.border = right_med_border_style
        ws.cell(row=start.row, column=start_col+7).border = right_med_border_style
        if cel.value < 0:
            cel.fill = PatternFill(fgColor="ff7878", fill_type="solid")
        elif cel.value > 1:
            cel.fill = PatternFill(fgColor="faaa01", fill_type="solid")
        start_col += 4

    for row in range(1, ws.max_row+1):
        for col in range(1, ws.max_column+1):
            ws.cell(row=row, column=col).alignment = alignment_style

    # saving file
    filename = "%s_grafik_na_%02d_%s_v_%s.xlsx" % (workplace.replace(" ", "_"), month, year, version)
    path = "app/xlsx_files/download/%s/%s/%s" % (year, month, workplace.replace(" ", "_"))
    if not os.path.exists(path):
        os.makedirs(path)
    wb.save("%s/%s" % (path, filename))
    return redirect(url_for("xlsx.download_schedule", year=year, month=month, workplace=workplace, filename=filename))


@bp.route("/xlsx_download/<year>/<month>/<workplace>/<path:filename>", methods=["GET", "POST"])
@login_required
def download_schedule(year, month, workplace, filename):
    """
    Allows to download file with schedule. Params creates path to file.
    :param year: schedule's year
    :param month: schedule's month
    :param workplace: schedule's workplace
    :param filename: name of file with schedule
    :return: downloading file
    """
    path = "/xlsx_files/download/%s/%s/%s/" % (year, month, workplace.replace(" ", "_"))
    to_file = current_app.root_path+path

    return send_from_directory(directory=to_file, filename=filename)


@bp.route("/upload-file", methods=["GET", "POST"])
@login_required
def upload_file():
    """
    Uploads chosen .xlsx file with schedule
    :return: template with form to choose schedule file
    :form validate return: redirection to function that creates dictionary based on uploaded file with schedule
    """
    year = request.args.get("year")
    month = request.args.get("month")
    workplace = request.args.get("workplace")
    hours = request.args.get("hours")
    title = "Grafiki - wybór pliku z grafikiem"
    form = UploadFile()
    if form.validate_on_submit():
        if "file" not in request.files:
            flash("Nie wybrano pliku")
            return redirect(request.url)

        file = request.files["file"]

        if file.filename == "":
            flash("Nie wybrano pliku")
            return redirect(request.url)

        if file and file.filename.rsplit('.', 1)[1].lower() == "xlsx":
            filename = secure_filename(file.filename)
            path = "app/xlsx_files/upload/"
            if not os.path.exists(path):
                os.makedirs(path)
            file.save("%s%s" % (path, filename))
            return redirect(url_for("xlsx.create_dict_from_file", path=path, filename=filename, year=year, month=month,
                                    workplace=workplace, hours=hours))
        else:
            flash("Nieprawidłowy rodzaj pliku. Plik musi być w formacie xlsx.")
            return redirect(request.url)

    return render_template("xlsx/upload_from_file.html", title=title, form=form)


@bp.route("/create_from_file", methods=["GET", "POST"])
@login_required
def create_dict_from_file():
    """
    Loads file with schedule and makes dictionary with data required to fulfill template
    :return: ?
    """
    def wrong_file(val):
        flash("W pliku nie znaleziono grafiku dla wybranego %s." % val)
        return redirect(url_for("xlsx.upload_file", month=month, year=year, workplace=workplace))

    path = request.args.get("path")
    filename = request.args.get("filename")
    file_path = path+filename
    year = int(request.args.get("year"))
    month = int(request.args.get("month"))
    workplace = request.args.get("workplace")
    hours = request.args.get("hours")
    db_workplace = Shop.query.filter_by(shopname=workplace).first()
    month_names = ["Styczeń", "Luty", "Marzec", "Kwiecień", "Maj", "Czerwiec", "Lipiec", "Sierpień",
                   "Wrzesień", "Październik", "Listopad", "Grudzień"]
    month_name = month_names[month-1]
    weekday_names = ["Poniedziałek", "Wtorek", "Środa", "Czwartek", "Piątek", "Sobota", "Niedziela"]
    cal = calendar.Calendar()
    prev = schedules.routes.prev_schedule(month, year, month_names, cal, workplace)
    wb = load_workbook(file_path, data_only=True)

    # finding sheet
    sheets = wb.sheetnames
    sh = None
    for sheet in sheets:
        sheet_month = sheet.lower().replace("grafik_", "").lower()
        if sheet_month.capitalize() == month_name:
            sh = wb[sheet]
    if sh is None:
        flash("W pliku nie znaleziono grafiku na wybrany miesiąc.")
        return redirect(url_for("xlsx.upload_file", month=month, year=year, workplace=workplace))

    # finding month cell
    month_cell = None
    for row in range(int(sh.max_row)):
        for col in range(int(sh.max_column)):
            current_cell = sh.cell(row=row+1, column=col+1)
            if type(current_cell.value) is str and current_cell.value.lower().replace(" ", "") == month_name.lower():
                month_cell = current_cell
    if month_cell is None:
        return wrong_file("miesiąca")

    # finding year
    year_cell = None
    for row in range(int(sh.max_row)):
        for col in range(int(sh.max_column)):
            current_cell = sh.cell(row=row+1, column=col+1)
            if type(current_cell.value) is int and current_cell.value == year:
                year_cell = current_cell
    if year_cell is None:
        return wrong_file("roku")

    # finding workplace
    workplace_cell = None
    for row in range(int(sh.max_row)):
        for col in range(int(sh.max_column)):
            current_cell = sh.cell(row=row + 1, column=col + 1)
            if type(current_cell.value) is str and current_cell.value.lower() == workplace.lower():
                workplace_cell = current_cell
    if workplace_cell is None:
        return wrong_file("sklepu")

    # finding workers
    workers = {}
    wrong_workers = []
    db_workers = []
    worker_row = workplace_cell.row
    worker_col = cifs(workplace_cell.column) + 2
    name_cell = sh.cell(row=worker_row, column=worker_col)
    surename_cell = sh.cell(row=worker_row+1, column=worker_col)
    while name_cell.value is not None and surename_cell.value is not None:
        name = name_cell.value.replace(" ", "") + " " + surename_cell.value.replace(" ", "")
        workers[name] = surename_cell
        #workers.append(name)
        worker_col += 4
        name_cell = sh.cell(row=worker_row, column=worker_col)
        surename_cell = sh.cell(row=worker_row + 1, column=worker_col)
    for worker in db_workplace.works.all():
        db_workers.append(str(worker))
    for worker in workers:
        if worker not in db_workers:
            wrong_workers.append(worker)
    if len(wrong_workers) > 0:
        flash("Sprawdź plik. Pracownicy nieprzypisani do sklepu: %s" % wrong_workers)

    # building dictionary with data for each day for each worker
    shdict = {}
    events = ("off", "in_work", "UW", "UNŻ", "L4", "UO", "UOJ", "UR", "UB")
    for worker in workers:
        for day in cal.itermonthdays(year, month):
            if day > 0:

                # begin hour
                begin = sh.cell(row=workers[worker].row+1+day, column=cifs(workers[worker].column)).value
                if begin is None:
                    begin = ""
                shdict["begin-%s-%d-%02d-%02d" % (worker.replace(" ", "_"), year, month, day)] = begin
                print("Po zmianie = " + worker)

                # end hour
                end = sh.cell(row=workers[worker].row+1+day, column=cifs(workers[worker].column)+1).value
                if end is None:
                    end = ""
                shdict["end-%s-%d-%02d-%02d" % (worker.replace(" ", "_"), year, month, day)] = end

                # event
                event = begin = sh.cell(row=workers[worker].row+1+day, column=cifs(workers[worker].column)+3).value
                if event is None:
                    event = "in_work"
                elif event == "X":
                    event = "off"
                if event not in events:
                    flash("Niewłaściwe oznaczenie zdarzenia w dniu %02d-%02d-%04d u pracownika %s." % (day, month, year,
                          worker))
                    return redirect(url_for("xlsx.upload_file", month=month, year=year, workplace=workplace))
                shdict["event-%s-%d-%02d-%02d" % (worker.replace(" ", "_"), year, month, day)] = event

    flash("Wczytałem grafik na %s %s" % (month_name.lower(), year))
    for i in shdict:
        print(i, shdict[i], type(shdict[i]))

    workers_to_schd = []
    for worker in workers:
        workers_to_schd.append(str(worker.replace(" ", "_")))
    print(workers_to_schd)

    return render_template("xlsx/empty_schedule.html", title="grafiki", workplace=workplace, year=year, month=month,
                           workers=workers_to_schd, month_name=month_name, wdn=weekday_names, cal=cal,
                           Billing_period=Billing_period, shdict=shdict, hours=hours, prev_shdict=prev["prev_shdict"],
                           prev_month=prev["month"], prev_hours=prev["hours"], prev_month_name=prev["month_name"],
                           prev_year=prev["year"], prev_workers=prev["workers"], workers_hours=prev["workers_hours"])
